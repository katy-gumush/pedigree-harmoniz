#!/usr/bin/env python3
"""Generate Test_Cases.xlsx (Name, Description, Steps, Expected result)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "Test_Cases.xlsx"

ROWS: list[tuple[str, str, str, str]] = [
    # ---------- Data layer (pytest: tests/test_data_level.py) ----------
    (
        "[Data] Clean CSV passes all integrity rules",
        "Validate that the baseline fixture passes every structural pedigree rule before any server starts.",
        "1. Load classpath fixture clean.csv.\n"
        "2. Run rule: required fields (ID, Name, Breed, Sex non-blank).\n"
        "3. Run rule: unique IDs across rows.\n"
        "4. Run rule: Sire_ID/Dam_ID empty or reference existing ID.\n"
        "5. Run rule: no dog lists itself as sire or dam.\n"
        "6. Run rule: no directed cycle following child→parent edges.\n"
        "7. Run rule: Height_cm and Weight_kg ≥ 0.",
        "All rules pass without AssertionError.",
    ),
    (
        "[Data] corrupt_bad_parent_id — invalid sire reference",
        "Detect a broken parent pointer (sire points to a dog id that does not exist).",
        "1. Load corrupt_bad_parent_id.csv.\n"
        "2. Confirm other rules still pass where applicable.\n"
        "3. Run valid-parent-reference rule.",
        "Assertion fails with message identifying the dog and unknown Sire_ID (e.g. 9999).",
    ),
    (
        "[Data] corrupt_duplicate_id — duplicate primary key",
        "Detect duplicate dog IDs which would corrupt lookups and silently overwrite rows when loaded into a map.",
        "1. Load corrupt_duplicate_id.csv.\n"
        "2. Run unique-ID rule.",
        "Assertion fails listing duplicate ID(s) and occurrence counts.",
    ),
    (
        "[Data] corrupt_immediate_loop — two-node mutual parent links",
        "Detect an impossible mutual parent relationship between two dogs.",
        "1. Load corrupt_immediate_loop.csv.\n"
        "2. Run directed-cycle detection on parent links.",
        "Assertion fails reporting a cycle on the ancestor walk.",
    ),
    (
        "[Data] corrupt_long_cycle — multi-node sire-only loop",
        "Detect a longer directed cycle in sire links (three or more dogs).",
        "1. Load corrupt_long_cycle.csv.\n"
        "2. Run directed-cycle detection on parent links.",
        "Assertion fails reporting a cycle during DFS of parent edges.",
    ),
    # ---------- API layer (pytest: tests/test_api.py) ----------
    (
        "[API] GET /api/dogs — list returns 200 and non-empty array",
        "Smoke-test the dogs collection endpoint shape and minimum size on full baseline data.",
        "1. GET /api/dogs with Accept: application/json.\n"
        "2. Inspect HTTP status and JSON body.",
        "Status 200; JSON array with size() > 500; first element has non-null id and non-empty name, breed, sex.",
    ),
    (
        "[API] GET /api/dogs — sampled rows contain required fields",
        "Ensure every sampled record exposes all fields required by the assignment for list views.",
        "1. GET /api/dogs.\n"
        "2. Assert list size > 500.\n"
        "3. Assert required keys on list[0], list[mid], list[last].\n"
        "4. Find entry with id=51 and assert required keys.",
        "Each checked object contains id, name, breed, sex, height_cm, weight_kg, sire_id, dam_id.",
    ),
    (
        "[API] GET /api/dogs/3 — founder Maggie matches CSV",
        "Verify detail endpoint for a founder dog with null parents.",
        "1. GET /api/dogs/3.",
        "Status 200; id=3; name=Maggie; breed=Boxer; sex=Male; sire_id and dam_id are null.",
    ),
    (
        "[API] GET /api/dogs/51 — Henry matches CSV parents",
        "Verify detail endpoint for a known dog with both parents.",
        "1. GET /api/dogs/51.",
        "Status 200; id=51; name=Henry; sire_id=1; dam_id=2.",
    ),
    (
        "[API] GET /api/dogs/{id} — unknown id returns 404",
        "API must surface missing resources with HTTP 404 and structured detail.",
        "1. GET /api/dogs/99999.",
        "Status 404; JSON detail equals \"Dog 99999 not found\".",
    ),
    (
        "[API] GET .../pedigree — response shape for dog 51",
        "Pedigree payload must include root plus ancestor and descendant collections.",
        "1. GET /api/dogs/51/pedigree.",
        "Status 200; body has root, ancestors, descendants; root.id = 51.",
    ),
    (
        "[API] Pedigree ancestors include both parents for dog 51",
        "Ancestor list must contain immediate parents Max (1) and Lucy (2).",
        "1. GET /api/dogs/51/pedigree.\n"
        "2. Collect ancestors.id.",
        "Ancestor id list contains 1 and 2.",
    ),
    (
        "[API] Ancestor depth capped at 5 generations",
        "Deep chains must not exceed MAX_GENERATIONS in API output.",
        "1. GET /api/dogs/452/pedigree.\n"
        "2. Read ancestors.depth values.",
        "Ancestors non-empty; max depth ≤ 5.",
    ),
    (
        "[API] Ancestor ids have no duplicates",
        "Same dog must not appear twice in the ancestor list for one root.",
        "1. GET /api/dogs/452/pedigree.\n"
        "2. Check ancestors.id for duplicates.",
        "No duplicate ids in ancestors list.",
    ),
    (
        "[API] Descendant ids have no duplicates (stress: founder Maggie)",
        "Large descendant trees must deduplicate nodes in the API list.",
        "1. GET /api/dogs/3/pedigree.\n"
        "2. Check descendants.id for duplicates.",
        "Descendants non-empty; no duplicate ids.",
    ),
    (
        "[API] Descendant depth capped at 5 generations",
        "Descendant generations in JSON must respect the same cap as ancestors.",
        "1. GET /api/dogs/3/pedigree.\n"
        "2. Read descendants.depth values.",
        "max(descendants.depth) ≤ 5.",
    ),
    (
        "[API] GET .../pedigree — unknown dog returns 404",
        "Pedigree for a non-existent id must not return 200 with empty lists.",
        "1. GET /api/dogs/99999/pedigree.",
        "Status 404; JSON detail equals \"Dog 99999 not found\".",
    ),
    (
        "[API] Founder dog has empty ancestors array",
        "Founder with no sire/dam in CSV must return an empty ancestors list.",
        "1. GET /api/dogs/3/pedigree.\n"
        "2. Read ancestors array length.",
        "ancestors is an empty JSON array.",
    ),
    (
        "[API] GET .../pedigree-network — local window coherent",
        "Pedigree network view returns nodes/edges within a bounded generation window.",
        "1. GET /api/dogs/51/pedigree-network?ancestors=2&descendants=2.\n"
        "2. Validate focus_id, nodes, edges; check node generations in [-2,2].\n"
        "3. Every edge references parent_id and child_id present in nodes.",
        "Status 200; focus_id=51; nodes.size between 1 and 200; each edge endpoints exist in nodes; generations within ±2.",
    ),
    (
        "[API] Invalid path id returns 422",
        "FastAPI path validation rejects non-integer dog ids.",
        "1. GET /api/dogs/not-a-number.",
        "Status 422; JSON contains detail describing validation error.",
    ),
    (
        "[API] GET /api/dataset — shape when switching enabled",
        "Dataset selection endpoint describes whether runtime switching is allowed.",
        "1. GET /api/dataset.",
        "Status 200; switching_enabled is true.",
    ),
    (
        "[API] GET /api/dataset — registry keys listed",
        "Options must expose all fixture keys used for corrupted-data demos.",
        "1. GET /api/dataset.\n"
        "2. Collect option keys.",
        "Options include full, clean, bad_parent, duplicate_id, immediate_loop, long_cycle.",
    ),
    (
        "[API] POST /api/dataset — unknown key returns 400",
        "Reject invalid dataset keys so API cannot enter an undefined state.",
        "1. POST /api/dataset with JSON {\"dataset\":\"no_such_key\"}.",
        "Status 400; detail \"Unknown dataset\".",
    ),
    (
        "[API] POST /api/dataset — cookie changes list size",
        "Selecting a miniature fixture reduces /api/dogs length; full restores baseline size.",
        "1. POST dataset=bad_parent; GET /api/dogs with returned cookies → expect 3 rows.\n"
        "2. POST dataset=duplicate_id; GET /api/dogs with cookie duplicate_id → expect 1 row.\n"
        "3. POST dataset=full; GET /api/dogs → expect size > 500.",
        "List sizes match selected fixture; cookie-backed selection persists across GET /api/dogs.",
    ),
    # ---------- UI layer (pytest: tests/ui + Playwright) ----------
    (
        "[UI] Home page shows dogs table",
        "Primary list view renders and exposes stable selectors for automation.",
        "1. Navigate to /. Wait for network idle.\n"
        "2. Locate [data-testid=dog-table].\n"
        "3. Count rows matching [data-testid^=dog-row-].",
        "dog-table visible; more than 100 dog rows rendered.",
    ),
    (
        "[UI] List contains known dog 51 (Henry)",
        "Regression anchor: a specific dog row is findable in the HTML table.",
        "1. Navigate to /.\n"
        "2. Find [data-testid=dog-row-51].\n"
        "3. Assert row text includes Henry.",
        "Row for dog 51 visible and contains name Henry.",
    ),
    (
        "[UI] Dog card shows all required birth-certificate fields",
        "Card view must display id, name, breed, sex, height, weight with non-blank values.",
        "1. Navigate to /dogs/51.\n"
        "2. Assert dog-card visible.\n"
        "3. For each test id dog-card-id,name,breed,sex,height,weight assert visible and non-blank.\n"
        "4. Assert dog-card-name equals Henry.",
        "All listed fields visible; name matches Henry.",
    ),
    (
        "[UI] Dog card shows correct sire and dam names",
        "Parent sections must surface the linked parent dog names from data.",
        "1. Navigate to /dogs/51.\n"
        "2. Read dog-card-sire and dog-card-dam text.",
        "Sire text contains Max; dam text contains Lucy.",
    ),
    (
        "[UI] Pedigree link navigates to pedigree URL",
        "Primary navigation from card to pedigree view works without manual URL typing.",
        "1. Navigate to /dogs/51.\n"
        "2. Click [data-testid=pedigree-link]. Wait for network idle.",
        "Browser URL ends with /dogs/51/pedigree.",
    ),
    (
        "[UI] Pedigree page shows correct root dog",
        "Pedigree header must identify the selected root by name and id.",
        "1. Navigate to /dogs/51/pedigree.\n"
        "2. Read [data-testid=pedigree-root] text.",
        "pedigree-root visible; text includes Henry and 51.",
    ),
    (
        "[UI] Pedigree lists both parents as ancestors",
        "Ancestor table rows must include parent ids 1 and 2 for dog 51.",
        "1. Navigate to /dogs/51/pedigree.\n"
        "2. Collect data-dog-id from [data-testid=pedigree-ancestor] rows.",
        "At least one ancestor row; ids include 1 and 2.",
    ),
    (
        "[UI] Pedigree ancestor rows have no duplicate dog ids (dog 452)",
        "UI must not render the same ancestor twice under depth cap and dedup logic.",
        "1. Navigate to /dogs/452/pedigree.\n"
        "2. Collect data-dog-id from ancestor rows.\n"
        "3. Compare count to distinct count.",
        "No duplicate data-dog-id values among ancestor rows.",
    ),
    (
        "[UI] Founder pedigree: no ancestor rows, fallback visible",
        "Founder dogs should show explicit empty-state instead of a broken table.",
        "1. Navigate to /dogs/3/pedigree.\n"
        "2. Count [data-testid=pedigree-ancestor] rows.\n"
        "3. Check [data-testid=pedigree-no-ancestors].",
        "Zero ancestor rows; pedigree-no-ancestors visible.",
    ),
    (
        "[UI] Founder pedigree shows descendants",
        "Descendant section populated for a dog with offspring in the tree.",
        "1. Navigate to /dogs/3/pedigree.\n"
        "2. Count [data-testid=pedigree-descendant] rows.",
        "At least one descendant row present.",
    ),
    (
        "[UI] Unknown dog shows user-safe error (no traceback)",
        "Graceful degradation: 404 dog must not expose server stack traces in HTML.",
        "1. Navigate to /dogs/99999.\n"
        "2. Assert [data-testid=error-message] visible.\n"
        "3. Assert body text does not contain Python \"Traceback\".",
        "Error message visible; page body does not include a Python traceback.",
    ),
    # ---------- Python unit tests (pytest: tests/test_data_level.py — pedigree helpers) ----------
    (
        "[Unit] build_pedigree ancestor chain",
        "Pure-function check: walking sire links collects ancestors with correct depths.",
        "1. Build in-memory dogs: 1 founder, 2 sire→1, 3 sire→2.\n"
        "2. Call build_pedigree(3, dogs).",
        "Root id 3; ancestor ids {1,2}; depths 2→1 and 1→2.",
    ),
    (
        "[Unit] build_pedigree descendants",
        "Descendant enumeration includes all direct children of the root.",
        "1. Build dogs: 1 founder; 2 and 3 sire→1.\n"
        "2. Call build_pedigree(1, dogs).",
        "Descendant ids {2,3}.",
    ),
    (
        "[Unit] Max depth clamp for ancestors",
        "When max_ancestors is set to 1, deeper ancestors are excluded.",
        "1. Chain 1←2←3←4.\n"
        "2. build_pedigree(4, dogs, max_ancestors=1, max_descendants=0).",
        "Ancestors contain only id 3; descendants empty.",
    ),
    (
        "[Unit] Missing parent id skipped",
        "Invalid sire reference does not crash; ancestors list stays empty.",
        "1. Single dog 2 with sire_id=999 (missing).\n"
        "2. build_pedigree(2, dogs).",
        "ancestors list is empty.",
    ),
    (
        "[Unit] initial_tree_depths — singleton dog",
        "Depth tuple for isolated node per app implementation.",
        "1. Dogs {1} only.\n"
        "2. initial_tree_depths(1, dogs).",
        "Returns (2, 2) per current pedigree_app implementation.",
    ),
    (
        "[Unit] initial_tree_depths — many direct children caps",
        "Stress case: many direct children affects initial depth estimate used by UI network.",
        "1. Root 1 with MANY_DIRECT_CHILDREN+1 children as sires only.\n"
        "2. initial_tree_depths(1, children).",
        "Returns (MAX_GENERATIONS, 2).",
    ),
    (
        "[Unit] initial_tree_depths — moderate branching",
        "Moderate fan-out and parents yield full window depth tuple.",
        "1. Dogs 1, 2←1, 3←1.\n"
        "2. initial_tree_depths(1, dogs).",
        "Returns (MAX_GENERATIONS, MAX_GENERATIONS).",
    ),
    (
        "[Unit] build_pedigree_network unknown focus",
        "Network builder returns empty nodes/edges for unknown focus id.",
        "1. build_pedigree_network(99, dogs, …) with dogs {1} only.",
        "nodes and edges empty.",
    ),
    (
        "[Unit] build_pedigree_network edges",
        "Edges reflect parent-child links within depth window.",
        "1. Dogs 1; 2 sire←1; 3 dam←2.\n"
        "2. build_pedigree_network(3, dogs, ancestor depth 5, descendant depth 5).",
        "Nodes {1,2,3}; edges include (1,2) and (2,3).",
    ),
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Test cases"

    headers = ("Name", "Description", "Steps", "Expected result")
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for name, desc, steps, expected in ROWS:
        ws.append([name, desc, steps, expected])

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = wrap

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 56
    ws.column_dimensions["D"].width = 48

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
